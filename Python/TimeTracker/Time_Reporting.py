# import openpyxl and tkinter modules - You will need to install them if not installed
from openpyxl import *
from tkinter import *

# other house keeping & QOL stuff
import getpass
username = getpass.getuser() # Generates a variable with the logged in username
from datetime import date
today = str(date.today()) # Generates a date string as YYYY-MM-DD
  
# globally declare wb and sheet variable 
# opening the existing excel file - Spreadsheet must be in the same folder as script or define the full path of using UNC
wb = load_workbook("time.xlsx") 
  
# create the sheet object 
sheet = wb.active 
  
def excel(): 
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 90
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 40
  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Date"
    sheet.cell(row=1, column=2).value = "Description"
    sheet.cell(row=1, column=3).value = "Time Spent"
    sheet.cell(row=1, column=4).value = "Service Now Reference"
    sheet.cell(row=1, column=5).value = "Work Type"
    sheet.cell(row=1, column=6).value = "Team Member"
  
# Function to set focus (cursor) 
# Function to set focus 
def focus1(event): 
    # set focus on the sem_field box 
    desc_field.focus_set() 
# Function to set focus 
def focus2(event): 
    # set focus on the form_no_field box 
    sn_no_field.focus_set() 
# Function to set focus 
def focus3(event): 
    # set focus on the contact_no_field box 
    work_type_field.focus_set() 
# Function to set focus 
def focus4(event): 
    # set focus on the contact_no_field box 
    user_field.focus_set() 

# Function for clearing the 
# contents of text entry boxes 
def clear(): 
    # clear the content of text entry box 
    desc_field.delete(0, END) 
    time_field.delete(0, END) 
    sn_no_field.delete(0, END)
    work_type_field.delete(0, END) 
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
    # if user not fill any entry 
    # then print "empty input" 
    if (desc_field.get() == "" and
        time_field.get() == "" and
        sn_no_field.get() == "" and
        work_type_field.get() == ""):
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = today 
        sheet.cell(row=current_row + 1, column=2).value = desc_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = time_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = sn_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = work_type_field.get()
        sheet.cell(row=current_row + 1, column=6).value = username 
  
        # save the file - Spreadsheet must be in the same folder as script or define the full path of using UNC
        wb.save("time.xlsx")
  
        # set focus on the name_field box 
        desc_field.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light blue') 
  
    # set the title of GUI window 
    root.title("Time/Status Reporting") 
  
    # set the configuration of GUI window 
    root.geometry("550x175") 
  
    excel() 
  
    # create a Form label 
    heading = Label(root, text="IT Time Reporting", bg="light blue") 
  
    # create a Description label 
    desc = Label(root, text="Description", bg="light blue") 
  
    # create a Time Worked label 
    time = Label(root, text="Time Worked", bg="light blue") 
  
    # create a Service Now Reference No lable 
    sn_no = Label(root, text="Service Now Reference No.", bg="light blue") 

    # create a Work Type lable 
    work_type = Label(root, text="Work Type.", bg="light blue")
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    desc.grid(row=1, column=0) 
    time.grid(row=2, column=0) 
    sn_no.grid(row=3, column=0) 
    work_type.grid(row=4, column=0) 
  
    # create a text entry box 
    # for typing the information 
    desc_field = Entry(root) 
    time_field = Entry(root) 
    sn_no_field = Entry(root)
    work_type_field = Entry(root)
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    desc_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    time_field.bind("<Return>", focus2) 

    # whenever the enter key is pressed 
    # then call the focus4 function 
    sn_no_field.bind("<Return>", focus3) 

    # whenever the enter key is pressed 
    # then call the focus3 function 
    work_type_field.bind("<Return>", focus4) 

    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    desc_field.grid(row=1, column=1, ipadx="100") 
    time_field.grid(row=2, column=1, ipadx="100") 
    sn_no_field.grid(row=3, column=1, ipadx="100")
    work_type_field.grid(row=4, column=1, ipadx="100")  
  
    # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="light gray", command=insert) 
    submit.grid(row=8, column=1) 
  
    # start the GUI 
    root.mainloop() 